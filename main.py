import streamlit as st
import google.generativeai as genai
import json
from typing import Dict, Any, List
import pandas as pd
import io
import tempfile
import os
import re
import base64
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from PyPDF2 import PdfReader
import docx
from datetime import datetime

api_key = st.secrets["GEMINI_API_KEY"]

class TextExtractor:    
    def __init__(self, api_key):
        genai.configure(api_key=api_key)
    
    def extract_text_from_image(self, image_content):
        try:
            image_base64 = base64.b64encode(image_content).decode('utf-8')
            
            prompt = """
            Ekstrak semua teks yang ada dalam gambar ini. Berikan teks lengkap persis seperti yang terlihat dalam dokumen, 
            pertahankan struktur dan format sebisa mungkin. Fokus pada keterbacaan dan akurasi.
            """
            
            model = genai.GenerativeModel('gemini-2.0-flash')
            response = model.generate_content([
                prompt,
                {
                    "mime_type": "image/jpeg",
                    "data": image_base64
                }
            ])
            
            return response.text
            
        except Exception as e:
            return f"Error extracting text from image: {str(e)}"

    def extract_text_from_pdf(self, pdf_content):
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
                tmp.write(pdf_content)
                tmp_path = tmp.name
            
            reader = PdfReader(tmp_path)
            text = ""
            for page in reader.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
            
            os.unlink(tmp_path)
            
            return text.strip() if text else "Tidak ada teks terbaca dari PDF"
            
        except Exception as e:
            return f"Gagal ekstrak teks dari PDF: {str(e)}"

    def extract_text_from_docx(self, docx_content):
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix=".docx") as tmp:
                tmp.write(docx_content)
                tmp_path = tmp.name
            
            doc = docx.Document(tmp_path)
            text = "\n".join([para.text for para in doc.paragraphs])
            
            os.unlink(tmp_path)
            
            return text.strip() if text else "Tidak ada teks terbaca dari Word"
            
        except Exception as e:
            return f"Gagal ekstrak teks dari Word: {str(e)}"

class JobRoleAnalyzer:
    
    def __init__(self, api_key):
        genai.configure(api_key=api_key)
    
    def generate_job_requirements_from_title(self, job_title: str, level: str) -> str:
        """Generate job requirements based on job title and level"""
        prompt = f"""
        Bayangkan Anda seorang talent acquisition / HR senior selama 30 tahun di berbagai jenis perusahaan. 
        Berikan saya common job requirements dan job descriptions untuk posisi {job_title} di level {level}.

        Buatkan job description yang komprehensif dan realistis yang mencakup:
        1. Job Overview/Summary
        2. Key Responsibilities (minimal 5-8 poin)  
        3. Required Skills (Technical, Soft Skills, Tools)
        4. Education Requirements
        5. Experience Requirements
        6. Preferred Qualifications/Nice to Have
        7. Salary Range (berikan estimasi yang realistis)
        8. Job Type (Full-time/Part-time/Contract)

        Gunakan bahasa Indonesia yang profesional dan sesuaikan dengan level {level}. 
        Berikan detail yang cukup untuk dapat dianalisis terhadap CV kandidat.
        """
        
        try:
            model = genai.GenerativeModel('gemini-2.0-flash')
            response = model.generate_content(prompt)
            return response.text
            
        except Exception as e:
            return f"Error generating job requirements: {str(e)}"
    
    def analyze_job_role(self, job_role_text: str, level: str) -> Dict[str, Any]:
        
        prompt = f"""
        Kamu adalah seorang HR expert dan job analyst yang berpengalaman di berbagai bidang perusahaan selama lebih dari 30 tahun.
        Analisis job role berikut untuk level {level}.

        **Job Role Content:**
        {job_role_text}

        **Level:** {level}

        Berikan analisis dalam format JSON dengan struktur berikut:

        {{
            "job_analysis": {{
                "title": "extracted or inferred job title",
                "company": "company name if mentioned",
                "level": "{level}",
                "location": "location if mentioned or 'Not specified'",
                "job_type": "Full-time/Part-time/Contract/Internship",
                "salary_range": "salary range if mentioned or 'Not specified'",
                "description": "brief job description summary"
            }},
            "required_skills": [
                {{
                    "skill_category": "Technical Skills",
                    "skills": ["skill1", "skill2", "skill3"]
                }},
                {{
                    "skill_category": "Soft Skills", 
                    "skills": ["skill1", "skill2", "skill3"]
                }},
                {{
                    "skill_category": "Tools & Technologies",
                    "skills": ["tool1", "tool2", "tool3"]
                }}
            ],
            "requirements": {{
                "education": "education requirements",
                "experience": "experience requirements", 
                "certifications": "certification requirements if any",
                "languages": "language requirements if any"
            }},
            "responsibilities": [
                "responsibility 1",
                "responsibility 2", 
                "responsibility 3"
            ],
            "nice_to_have": [
                "nice to have skill/experience 1",
                "nice to have skill/experience 2"
            ]
        }}

        Berikan HANYA response JSON, tanpa tambahan teks atau formatting lain.
        Gunakan bahasa Indonesia untuk deskripsi dan bahasa Inggris untuk skill names.
        """
        
        try:
            model = genai.GenerativeModel('gemini-2.0-flash')
            response = model.generate_content(prompt)
            
            response_text = response.text.strip()
            
            if response_text.startswith('```json'):
                response_text = response_text.replace('```json', '').replace('```', '').strip()
            
            json_response = json.loads(response_text)
            return json_response
            
        except json.JSONDecodeError:
            try:
                start_idx = response_text.find('{')
                end_idx = response_text.rfind('}') + 1
                if start_idx != -1 and end_idx != -1:
                    json_str = response_text[start_idx:end_idx]
                    json_response = json.loads(json_str)
                    return json_response
                else:
                    st.error("Could not extract valid JSON from AI response")
                    return None
            except Exception as e:
                st.error(f"Failed to parse JSON response: {str(e)}")
                return None
        
        except Exception as e:
            st.error(f"Error analyzing job role: {str(e)}")
            return None

class CVAnalyzer:
    def __init__(self, api_key):
        genai.configure(api_key=api_key)
    
    def analyze_cv_against_job(self, cv_text: str, job_analysis: Dict[str, Any]) -> Dict[str, Any]:        
        
        job_requirements = json.dumps(job_analysis, indent=2, ensure_ascii=False)
        
        prompt = f"""
        Kamu adalah seorang HR expert dan career consultant yang berpengalaman selama lebih dari 30 tahun di berbagai jenis perusahaan.
        Analisis CV berikut terhadap job requirements yang spesifik.

        **CV Content:**
        {cv_text}

        **Job Requirements:**
        {job_requirements}

        Analisis CV terhadap job requirements di atas dan berikan penilaian dalam format JSON dengan struktur berikut:

        {{
            "resume_evaluation_matrix": [
                {{
                    "kriteria": "Clarity & Structure",
                    "skor": <1-10>,
                    "catatan": "penjelasan detail"
                }},
                {{
                    "kriteria": "Relevance to Target Role", 
                    "skor": <1-10>,
                    "catatan": "penjelasan detail seberapa relevan dengan job requirements"
                }},
                {{
                    "kriteria": "Impact & Achievement",
                    "skor": <1-10>,
                    "catatan": "penjelasan detail"
                }},
                {{
                    "kriteria": "Language & Grammar",
                    "skor": <1-10>,
                    "catatan": "penjelasan detail"
                }},
                {{
                    "kriteria": "ATS Compatibility",
                    "skor": <1-10>,
                    "catatan": "penjelasan detail"
                }},
                {{
                    "kriteria": "Design & Visual Appeal",
                    "skor": <1-10>,
                    "catatan": "penjelasan detail"
                }},
                {{
                    "kriteria": "Length Appropriateness",
                    "skor": <1-10>,
                    "catatan": "penjelasan detail"
                }},
                {{
                    "kriteria": "Job Requirements Match",
                    "skor": <1-10>,
                    "catatan": "penjelasan seberapa match dengan requirements job yang diupload"
                }}
            ],
            "skill_gap_analysis": [
                {{
                    "skill_area": "nama area skill dari job requirements",
                    "required_skill": "detail skill yang dibutuhkan dari job posting",
                    "found_in_resume": "yes/no/partial",
                    "skill_gap": <0-5>,
                    "catatan": "penjelasan gap dan saran improvement"
                }}
            ],
            "job_readiness_index": [
                {{
                    "aspect": "Relevansi Pengalaman Kerja",
                    "skor": <1-10>,
                    "catatan": "penjelasan relevansi pengalaman dengan job requirements"
                }},
                {{
                    "aspect": "Kesesuaian Skill",
                    "skor": <1-10>, 
                    "catatan": "penjelasan kesesuaian skill dengan yang diminta di job posting"
                }},
                {{
                    "aspect": "Education & Certification Match",
                    "skor": <1-10>,
                    "catatan": "penjelasan kesesuaian pendidikan dan sertifikasi"  
                }},
                {{
                    "aspect": "Resume Readiness",
                    "skor": <1-10>,
                    "catatan": "penjelasan kesiapan resume untuk apply job ini"
                }},
                {{
                    "aspect": "Interview Preparedness", 
                    "skor": <1-10>,
                    "catatan": "penjelasan kesiapan untuk interview berdasarkan job requirements"
                }}
            ]
        }}

        Berikan HANYA response JSON, tanpa tambahan teks atau formatting lain.
        Gunakan bahasa Indonesia yang mudah dipahami dan berikan penilaian yang objektif dan konstruktif.
        Fokus pada kesesuaian antara CV dengan job requirements yang spesifik.
        """
        
        try:
            model = genai.GenerativeModel('gemini-2.0-flash')
            response = model.generate_content(prompt)
            
            response_text = response.text.strip()
            
            if response_text.startswith('```json'):
                response_text = response_text.replace('```json', '').replace('```', '').strip()
            
            json_response = json.loads(response_text)
            return json_response
            
        except json.JSONDecodeError:
            try:
                start_idx = response_text.find('{')
                end_idx = response_text.rfind('}') + 1
                if start_idx != -1 and end_idx != -1:
                    json_str = response_text[start_idx:end_idx]
                    json_response = json.loads(json_str)
                    return json_response
                else:
                    st.error("Could not extract valid JSON from AI response")
                    return None
            except Exception as e:
                st.error(f"Failed to parse JSON response: {str(e)}")
                return None
        
        except Exception as e:
            st.error(f"Error analyzing CV: {str(e)}")
            return None

class ExcelReportGenerator:
    
    @staticmethod
    def create_formatted_report(analysis_result: Dict, job_analysis: Dict, candidate_name: str):
        
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            
            if 'resume_evaluation_matrix' in analysis_result:
                rem_data = []
                total_score = 0
                
                for item in analysis_result['resume_evaluation_matrix']:
                    score = item.get('skor', 0)
                    total_score += score
                    rem_data.append({
                        'Kriteria': item.get('kriteria', ''),
                        'Skor (1-10)': score,
                        'Catatan': item.get('catatan', '')
                    })
                
                percentage = (total_score / 80) * 100
                if percentage < 50:
                    summary_note = "Resume perlu revisi besar-besaran, perhatikan semua aspek"
                elif percentage < 65:
                    summary_note = "Sudah lumayan, namun perlu beberapa perbaikan"
                elif percentage < 80:
                    summary_note = "Resume berkualitas baik"
                else:
                    summary_note = "Resume sangat profesional dan menarik"
                
                rem_data.append({
                    'Kriteria': 'TOTAL SCORE',
                    'Skor (1-10)': total_score,
                    'Catatan': f"PG Score: {percentage:.1f}% - {summary_note}"
                })
                
                rem_data.append({
                    'Kriteria': '',
                    'Skor (1-10)': '',
                    'Catatan': f"PG Score: {percentage:.1f}%"
                })
                
                df_rem = pd.DataFrame(rem_data)
                df_rem.to_excel(writer, sheet_name='Resume Evaluation Matrix', index=False)
            
            job_data = []
            
            if 'job_analysis' in job_analysis:
                job_info = job_analysis['job_analysis']
                job_data.append({'Section': 'Basic Info', 'Details': f"Title: {job_info.get('title', '')}"})
                job_data.append({'Section': 'Basic Info', 'Details': f"Company: {job_info.get('company', '')}"})
                job_data.append({'Section': 'Basic Info', 'Details': f"Level: {job_info.get('level', '')}"})
                job_data.append({'Section': 'Basic Info', 'Details': f"Location: {job_info.get('location', '')}"})
                job_data.append({'Section': 'Basic Info', 'Details': f"Type: {job_info.get('job_type', '')}"})
                job_data.append({'Section': 'Basic Info', 'Details': f"Salary: {job_info.get('salary_range', '')}"})
            
            if 'required_skills' in job_analysis:
                for skill_cat in job_analysis['required_skills']:
                    skills_text = ", ".join(skill_cat.get('skills', []))
                    job_data.append({
                        'Section': f"Required Skills - {skill_cat.get('skill_category', '')}",
                        'Details': skills_text
                    })
            
            if 'requirements' in job_analysis:
                req = job_analysis['requirements']
                for key, value in req.items():
                    job_data.append({
                        'Section': f"Requirements - {key.title()}",
                        'Details': value
                    })
            
            if 'responsibilities' in job_analysis:
                for i, resp in enumerate(job_analysis['responsibilities'], 1):
                    job_data.append({
                        'Section': f"Responsibility {i}",
                        'Details': resp
                    })
            
            df_job = pd.DataFrame(job_data)
            df_job.to_excel(writer, sheet_name='Job Requirements', index=False)
            
            if 'skill_gap_analysis' in analysis_result:
                sgi_data = []
                total_gap = 0
                item_count = 0
                
                for item in analysis_result['skill_gap_analysis']:
                    gap_score = item.get('skill_gap', 0)
                    total_gap += gap_score
                    item_count += 1
                    sgi_data.append({
                        'Skill Area': item.get('skill_area', ''),
                        'Required Skill': item.get('required_skill', ''),
                        'Found in Resume?': item.get('found_in_resume', ''),
                        'Skill Gap (0-5)': gap_score,
                        'Catatan': item.get('catatan', '')
                    })
                
                if item_count > 0:
                    avg_gap = total_gap / item_count
                    gap_percentage = max(0, 100 - (avg_gap / 5) * 100)
                    
                    if gap_percentage < 40:
                        gap_note = "Banyak skill yang belum dimiliki, perlu training signifikan"
                    elif gap_percentage < 60:
                        gap_note = "Beberapa skill perlu ditingkatkan"
                    elif gap_percentage < 80:
                        gap_note = "Skill sudah cukup baik, perlu sedikit improvement"
                    else:
                        gap_note = "Skill sangat sesuai dengan requirements"
                    
                    sgi_data.append({
                        'Skill Area': 'TOTAL ANALYSIS',
                        'Required Skill': '',
                        'Found in Resume?': '',
                        'Skill Gap (0-5)': f"Avg: {avg_gap:.1f}",
                        'Catatan': f"PG Score: {gap_percentage:.1f}% - {gap_note}"
                    })
                    
                    sgi_data.append({
                        'Skill Area': '',
                        'Required Skill': '',
                        'Found in Resume?': '',
                        'Skill Gap (0-5)': '',
                        'Catatan': f"PG Score: {gap_percentage:.1f}%"
                    })
                
                df_sgi = pd.DataFrame(sgi_data)
                df_sgi.to_excel(writer, sheet_name='Skill Gap Analysis', index=False)
            
            if 'job_readiness_index' in analysis_result:
                jri_data = []
                total_score = 0
                
                for item in analysis_result['job_readiness_index']:
                    score = item.get('skor', 0)
                    total_score += score
                    jri_data.append({
                        'Aspect': item.get('aspect', ''),
                        'Score (1-10)': score,
                        'Catatan': item.get('catatan', '')
                    })
                
                percentage = (total_score / 50) * 100
                if percentage < 40:
                    summary_note = "Perlu persiapan signifikan di semua aspek"
                elif percentage < 60:
                    summary_note = "Perlu persiapan lebih di kesesuaian skill dan pengalaman"
                elif percentage < 80:
                    summary_note = "Cukup siap, perlu perbaikan minor"
                else:
                    summary_note = "Sangat siap untuk posisi yang dilamar"
                
                jri_data.append({
                    'Aspect': 'TOTAL SCORE',
                    'Score (1-10)': total_score,
                    'Catatan': f"PG Score: {percentage:.1f}% - {summary_note}"
                })
                
                jri_data.append({
                    'Aspect': '',
                    'Score (1-10)': '',
                    'Catatan': f"PG Score: {percentage:.1f}%"
                })
                
                df_jri = pd.DataFrame(jri_data)
                df_jri.to_excel(writer, sheet_name='Job Readiness Index', index=False)
        
        output.seek(0)
        wb = load_workbook(output)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            ws.insert_rows(1)
            ws['A1'] = sheet_name
            ws['A1'].font = Font(bold=True, size=16)
            ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
            
            last_column = chr(64 + len(ws[2])) if len(ws[2]) > 0 else 'A'
            ws.merge_cells(f'A1:{last_column}1')
            
            for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
                for cell in row:
                    if cell.column == 2 and isinstance(cell.value, (int, float)):  # Score columns
                        if cell.value < 5:
                            cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                            cell.font = Font(color="FFFFFF", bold=True)
            
            for col_idx, column in enumerate(ws.columns, 1):
                max_length = 0
                column_letter = chr(64 + col_idx)
                
                for cell in column:
                    try:
                        if hasattr(cell, 'value') and cell.value is not None:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        formatted_output = io.BytesIO()
        wb.save(formatted_output)
        formatted_output.seek(0)
        
        return formatted_output

def extract_text_from_file(uploaded_file, extractor: TextExtractor) -> str:
    try:
        file_content = uploaded_file.read()
        
        if uploaded_file.type == "application/pdf":
            return extractor.extract_text_from_pdf(file_content)
        elif uploaded_file.type in ["application/vnd.openxmlformats-officedocument.wordprocessingml.document", 
                                   "application/msword"]:
            return extractor.extract_text_from_docx(file_content)
        else:
            return extractor.extract_text_from_image(file_content)
    except Exception as e:
        return f"Error extracting text: {str(e)}"

def main():
    """Main Streamlit app"""
    
    st.set_page_config(
        page_title="Resume Analyzer",
        page_icon="📄",
        layout="wide"
    )

    st.title("Resume Analyzer")
    
    if 'cv_text' not in st.session_state:
        st.session_state.cv_text = ""
    if 'job_role_text' not in st.session_state:
        st.session_state.job_role_text = ""
    if 'analysis_result' not in st.session_state:
        st.session_state.analysis_result = None
    if 'job_analysis' not in st.session_state:
        st.session_state.job_analysis = None
    if 'candidate_name' not in st.session_state:
        st.session_state.candidate_name = ""
    if 'edited_analysis' not in st.session_state:
        st.session_state.edited_analysis = None
    if 'job_title' not in st.session_state:
        st.session_state.job_title = ""
    
    extractor = TextExtractor(api_key)
    cv_analyzer = CVAnalyzer(api_key)
    job_analyzer = JobRoleAnalyzer(api_key)
    
    st.header("Input Section")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("Upload CV")
        cv_file = st.file_uploader(
            "Upload CV (PDF/Word/Image):",
            type=['pdf', 'docx', 'doc', 'png', 'jpg', 'jpeg'],
            help="Upload CV dalam format PDF, Word document, atau gambar",
            key="cv_upload"
        )
        
        if cv_file is not None:
            with st.spinner("Extracting text from CV..."):
                cv_text = extract_text_from_file(cv_file, extractor)
                st.session_state.cv_text = cv_text
        
        if st.session_state.cv_text:
            st.subheader("Extracted CV Text (Editable)")
            edited_cv_text = st.text_area(
                "Edit CV text if needed:",
                value=st.session_state.cv_text,
                height=200,
                help="You can edit the extracted text to improve accuracy",
                key="cv_text_area"
            )
            st.session_state.cv_text = edited_cv_text
    
    with col2:
        st.subheader("Job Role Requirements")
        
        job_input_method = st.radio(
            "How would you like to provide job role information?",
            options=["Upload File (PDF/Word/Image)", "Paste Text", "Generate from Job Title"],
            key="job_input_method"
        )
        
        if job_input_method == "Upload File (PDF/Word/Image)":
            job_file = st.file_uploader(
                "Upload Job Role (PDF/Word/Image):",
                type=['pdf', 'docx', 'doc', 'png', 'jpg', 'jpeg'],
                help="Upload job description dalam format PDF, Word document, atau gambar",
                key="job_upload"
            )
            
            if job_file is not None:
                with st.spinner("Extracting text from job role..."):
                    job_text = extract_text_from_file(job_file, extractor)
                    st.session_state.job_role_text = job_text
        
        elif job_input_method == "Paste Text":
            job_text_input = st.text_area(
                "Paste Job Role Description:",
                height=200,
                placeholder="Paste the complete job description here...",
                key="job_text_input"
            )
            st.session_state.job_role_text = job_text_input
        
        else:  # Generate from Job Title
            job_title_input = st.text_input(
                "Job Title:",
                placeholder="e.g., Software Engineer, Marketing Manager, Data Analyst",
                help="Enter the job title you want to analyze",
                key="job_title_input"
            )
            st.session_state.job_title = job_title_input
            
            if job_title_input:
                st.info("Job requirements will be generated automatically based on the job title when you click 'Analyze CV Against Job Role'")
        
        # Show editable job role text if it exists
        if st.session_state.job_role_text and job_input_method != "Generate from Job Title":
            st.subheader("Job Role Text (Editable)")
            edited_job_text = st.text_area(
                "Edit job role text if needed:",
                value=st.session_state.job_role_text,
                height=200,
                help="You can edit the extracted text to improve accuracy",
                key="job_text_area"
            )
            st.session_state.job_role_text = edited_job_text
    
    st.subheader("Additional Information:")
    col1, col2 = st.columns(2)
    
    with col1:
        candidate_name = st.text_input("Nama Kandidat:", value=st.session_state.candidate_name)
        st.session_state.candidate_name = candidate_name
    
    with col2:
        level = st.selectbox(
            "Job Level:",
            options=["entry", "middle", "senior", "internship"],
            index=0,
            key="job_level"
        )
    
    # Determine if analysis can be performed
    job_input_method = st.session_state.get('job_input_method', 'Upload File (PDF/Word/Image)')
    can_analyze = False
    
    if job_input_method == "Generate from Job Title":
        can_analyze = st.session_state.cv_text and st.session_state.job_title
    else:
        can_analyze = st.session_state.cv_text and st.session_state.job_role_text
    
    if st.button("Analyze CV Against Job Role", type="primary", disabled=not can_analyze):
        if st.session_state.cv_text:
            with st.spinner("Analyzing job role and comparing with CV..."):
                
                # Generate job requirements if using job title method
                if job_input_method == "Generate from Job Title" and st.session_state.job_title:
                    with st.spinner("Generating job requirements from job title..."):
                        generated_job_text = job_analyzer.generate_job_requirements_from_title(
                            st.session_state.job_title, level
                        )
                        st.session_state.job_role_text = generated_job_text
                        
                        # Show the generated job requirements
                        with st.expander("📋 Generated Job Requirements (Click to view)", expanded=False):
                            st.write(generated_job_text)
                
                # Analyze the job role (whether uploaded/pasted or generated)
                job_analysis = job_analyzer.analyze_job_role(
                    st.session_state.job_role_text,
                    level
                )
                st.session_state.job_analysis = job_analysis

                # Analyze CV against job role
                analysis_result = cv_analyzer.analyze_cv_against_job(
                    st.session_state.cv_text,
                    job_analysis
                )
                st.session_state.analysis_result = analysis_result

                if analysis_result:
                    st.success("Analysis completed successfully ✅")

    # Display results if available
    if st.session_state.analysis_result and st.session_state.job_analysis:
        st.header("Analysis Results")

        # Resume Evaluation Matrix
        if "resume_evaluation_matrix" in st.session_state.analysis_result:
            st.subheader("📊 Resume Evaluation Matrix")
            df_rem = pd.DataFrame(st.session_state.analysis_result["resume_evaluation_matrix"])
            st.dataframe(df_rem, use_container_width=True)

        # Skill Gap Analysis
        if "skill_gap_analysis" in st.session_state.analysis_result:
            st.subheader("🛠️ Skill Gap Analysis")
            df_sga = pd.DataFrame(st.session_state.analysis_result["skill_gap_analysis"])
            st.dataframe(df_sga, use_container_width=True)

        # Job Readiness Index
        if "job_readiness_index" in st.session_state.analysis_result:
            st.subheader("🚀 Job Readiness Index")
            df_jri = pd.DataFrame(st.session_state.analysis_result["job_readiness_index"])
            st.dataframe(df_jri, use_container_width=True)

        # Download Excel Report
        st.subheader("📥 Download Report")
        if st.button("Generate Excel Report"):
            report_file = ExcelReportGenerator.create_formatted_report(
                st.session_state.analysis_result,
                st.session_state.job_analysis,
                st.session_state.candidate_name
            )

            st.download_button(
                label="⬇️ Download Excel Report",
                data=report_file,
                file_name=f"CV_Analysis_Report_{st.session_state.candidate_name}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )


if __name__ == "__main__":
    main()
